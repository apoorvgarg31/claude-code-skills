#!/usr/bin/env python3
"""
PDF Data Extractor - Extract structured data from PDFs.

Usage:
    python extract.py document.pdf [--format json|csv|xlsx] [--output file] [--tables-only] [--page N]
"""

import sys
import json
import argparse
import csv
from pathlib import Path
from io import StringIO

try:
    import fitz  # PyMuPDF
except ImportError:
    print("Error: PyMuPDF not installed. Run: pip install pymupdf", file=sys.stderr)
    sys.exit(1)


def extract_metadata(doc: fitz.Document) -> dict:
    """Extract document metadata."""
    meta = doc.metadata or {}
    return {
        "title": meta.get("title", ""),
        "author": meta.get("author", ""),
        "subject": meta.get("subject", ""),
        "creator": meta.get("creator", ""),
        "producer": meta.get("producer", ""),
        "creation_date": meta.get("creationDate", ""),
        "modification_date": meta.get("modDate", ""),
        "page_count": doc.page_count,
        "is_encrypted": doc.is_encrypted,
    }


def extract_tables(page: fitz.Page) -> list:
    """Extract tables from a PDF page using PyMuPDF's table finder."""
    tables = []
    try:
        # PyMuPDF 1.23+ has find_tables()
        tab_finder = page.find_tables()
        for table in tab_finder:
            extracted = table.extract()
            if extracted:
                tables.append(extracted)
    except AttributeError:
        # Older PyMuPDF versions - fall back to basic extraction
        pass
    return tables


def extract_text(page: fitz.Page, preserve_layout: bool = True) -> str:
    """Extract text from a PDF page."""
    if preserve_layout:
        return page.get_text("text")
    return page.get_text()


def extract_form_fields(doc: fitz.Document) -> list:
    """Extract form fields from the PDF."""
    fields = []
    try:
        for page in doc:
            widgets = page.widgets()
            if widgets:
                for widget in widgets:
                    fields.append({
                        "page": page.number + 1,
                        "field_name": widget.field_name or "",
                        "field_type": widget.field_type_string,
                        "field_value": widget.field_value or "",
                        "rect": list(widget.rect),
                    })
    except Exception:
        pass
    return fields


def process_pdf(pdf_path: str, page_num: int = None, tables_only: bool = False) -> dict:
    """Process PDF and extract all data."""
    doc = fitz.open(pdf_path)
    
    if doc.is_encrypted:
        raise ValueError(f"PDF is encrypted and cannot be processed: {pdf_path}")
    
    result = {
        "source": str(pdf_path),
        "metadata": extract_metadata(doc),
        "pages": [],
        "form_fields": extract_form_fields(doc),
    }
    
    # Determine which pages to process
    if page_num is not None:
        if page_num < 1 or page_num > doc.page_count:
            raise ValueError(f"Page {page_num} out of range (1-{doc.page_count})")
        pages_to_process = [doc[page_num - 1]]
    else:
        pages_to_process = doc
    
    for page in pages_to_process:
        page_data = {
            "number": page.number + 1,
            "tables": extract_tables(page),
        }
        
        if not tables_only:
            page_data["text"] = extract_text(page)
        
        result["pages"].append(page_data)
    
    doc.close()
    return result


def output_json(data: dict, output_file: str = None):
    """Output data as JSON."""
    json_str = json.dumps(data, indent=2, ensure_ascii=False)
    if output_file:
        Path(output_file).write_text(json_str)
        print(f"JSON written to: {output_file}", file=sys.stderr)
    else:
        print(json_str)


def output_csv(data: dict, output_file: str = None):
    """Output tables as CSV."""
    output = StringIO()
    writer = csv.writer(output)
    
    table_count = 0
    for page in data["pages"]:
        for table in page.get("tables", []):
            if table_count > 0:
                writer.writerow([])  # Blank line between tables
            writer.writerow([f"# Table {table_count + 1} (Page {page['number']})"])
            for row in table:
                # Clean None values
                clean_row = [cell if cell is not None else "" for cell in row]
                writer.writerow(clean_row)
            table_count += 1
    
    csv_content = output.getvalue()
    
    if output_file:
        Path(output_file).write_text(csv_content)
        print(f"CSV written to: {output_file}", file=sys.stderr)
    else:
        print(csv_content)


def output_xlsx(data: dict, output_file: str):
    """Output data as Excel workbook."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        print("Error: openpyxl not installed. Run: pip install openpyxl", file=sys.stderr)
        sys.exit(1)
    
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Header style
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # Text sheet
    text_ws = wb.create_sheet("Text")
    text_ws.append(["Page", "Content"])
    text_ws["A1"].font = header_font
    text_ws["B1"].font = header_font
    
    for page in data["pages"]:
        text_content = page.get("text", "")
        text_ws.append([page["number"], text_content[:32000]])  # Excel cell limit
    
    # Tables sheets
    table_num = 1
    for page in data["pages"]:
        for table in page.get("tables", []):
            ws = wb.create_sheet(f"Table_{table_num}")
            for row_idx, row in enumerate(table):
                clean_row = [cell if cell is not None else "" for cell in row]
                ws.append(clean_row)
                if row_idx == 0:
                    for cell in ws[1]:
                        cell.font = header_font
                        cell.fill = header_fill
            table_num += 1
    
    # Form fields sheet
    if data.get("form_fields"):
        ff_ws = wb.create_sheet("Form_Fields")
        ff_ws.append(["Page", "Field Name", "Type", "Value"])
        for cell in ff_ws[1]:
            cell.font = header_font
        for field in data["form_fields"]:
            ff_ws.append([
                field["page"],
                field["field_name"],
                field["field_type"],
                field["field_value"],
            ])
    
    # Metadata sheet
    meta_ws = wb.create_sheet("Metadata")
    meta_ws.append(["Property", "Value"])
    meta_ws["A1"].font = header_font
    meta_ws["B1"].font = header_font
    for key, value in data["metadata"].items():
        meta_ws.append([key, str(value)])
    
    if not output_file:
        output_file = Path(data["source"]).stem + ".xlsx"
    
    wb.save(output_file)
    print(f"Excel written to: {output_file}", file=sys.stderr)


def main():
    parser = argparse.ArgumentParser(
        description="Extract structured data from PDFs",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
    python extract.py document.pdf
    python extract.py report.pdf --format xlsx --output data.xlsx
    python extract.py forms.pdf --format csv --tables-only
    python extract.py large.pdf --page 5 --format json
        """
    )
    
    parser.add_argument("pdf_file", help="Path to PDF file")
    parser.add_argument(
        "--format", "-f",
        choices=["json", "csv", "xlsx"],
        default="json",
        help="Output format (default: json)"
    )
    parser.add_argument(
        "--output", "-o",
        help="Output file path (required for xlsx, optional for others)"
    )
    parser.add_argument(
        "--tables-only", "-t",
        action="store_true",
        help="Extract only tables, skip text content"
    )
    parser.add_argument(
        "--page", "-p",
        type=int,
        help="Extract specific page only (1-indexed)"
    )
    
    args = parser.parse_args()
    
    # Validate PDF file exists
    pdf_path = Path(args.pdf_file)
    if not pdf_path.exists():
        print(f"Error: File not found: {args.pdf_file}", file=sys.stderr)
        sys.exit(1)
    
    if not pdf_path.suffix.lower() == ".pdf":
        print(f"Warning: File may not be a PDF: {args.pdf_file}", file=sys.stderr)
    
    # xlsx requires output file
    if args.format == "xlsx" and not args.output:
        args.output = pdf_path.stem + ".xlsx"
    
    try:
        data = process_pdf(
            args.pdf_file,
            page_num=args.page,
            tables_only=args.tables_only
        )
        
        if args.format == "json":
            output_json(data, args.output)
        elif args.format == "csv":
            output_csv(data, args.output)
        elif args.format == "xlsx":
            output_xlsx(data, args.output)
            
    except ValueError as e:
        print(f"Error: {e}", file=sys.stderr)
        sys.exit(1)
    except Exception as e:
        print(f"Error processing PDF: {e}", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
